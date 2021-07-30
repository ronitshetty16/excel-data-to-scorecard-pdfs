from win32com import client
from fpdf import FPDF

import openpyxl


# Give the location of the file
path = "Dummy Data.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj1 = sheet_obj.cell(row=2, column=2)
cell_obj2 = sheet_obj.cell(row=3, column=2)
cell_obj3 = sheet_obj.cell(row=2, column=3)
cell_obj4 = sheet_obj.cell(row=3, column=3)
cell_obj5 = sheet_obj.cell(row=2, column=4)
cell_obj6 = sheet_obj.cell(row=3, column=4)
cell_obj7 = sheet_obj.cell(row=2, column=5)
cell_obj8 = sheet_obj.cell(row=3, column=5)
cell_obj9 = sheet_obj.cell(row=2, column=6)
cell_obj10 = sheet_obj.cell(row=3, column=6)
cell_obj11 = sheet_obj.cell(row=2, column=7)
cell_obj12 = sheet_obj.cell(row=3, column=7)
cell_obj13 = sheet_obj.cell(row=2, column=8)
cell_obj14 = sheet_obj.cell(row=3, column=8)
cell_obj15 = sheet_obj.cell(row=2, column=9)
cell_obj16 = sheet_obj.cell(row=3, column=9)
cell_obj17 = sheet_obj.cell(row=2, column=10)
cell_obj18 = sheet_obj.cell(row=3, column=10)
cell_obj19 = sheet_obj.cell(row=2, column=11)
cell_obj20 = sheet_obj.cell(row=3, column=11)
cell_obj21 = sheet_obj.cell(row=2, column=12)
cell_obj22 = sheet_obj.cell(row=3, column=12)
cell_obj23 = sheet_obj.cell(row=2, column=13)
cell_obj24 = sheet_obj.cell(row=3, column=13)
cell_object = sheet_obj.cell(row=2, column=20)
cell_objr = sheet_obj.cell(row=3, column=20)
# Print value of cell object
# using the value attribute


class PDF(FPDF):
    def header(self):
        # Logo
        self.image('Logo.jpg', 10, 8, 25)
        # font
        self.set_font('helvetica', 'I', 20)
        # Padding
        self.cell(60)
        # Title
        self.cell(100, 10, 'General Aptitude Examination',
                  border=True, ln=1, align='C')
        # Line break
        self.ln(20)


# create FPDF object
# Layout ('P','L')
# Unit ('mm', 'cm', 'in')
# format ('A3', 'A4' (default), 'A5', 'Letter', 'Legal', (100,150))
pdf = PDF('P', 'mm', 'Letter')

# Add a page
pdf.add_page()

# specify font
# fonts ('times', 'courier', 'helvetica', 'symbol', 'zpfdingbats')
# 'B' (bold), 'U' (underline), 'I' (italics), '' (regular), combination (i.e., ('BU'))
pdf.set_font('helvetica', 'B', 12)

# Add text
# w = width
# h = height
# txt = your text
# ln (0 False; 1 True - move cursor down to next line)
# border (0 False; 1 True - add border around cell)
pdf.cell(20, 7, '%s :' % (cell_obj1.value))
pdf.image('Pics for assignment/ABC1 XYZ1.png', 145, 40, 60)
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj2.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj3.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj5.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj7.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value) + ' %s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj9.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj10.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj11.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj12.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj13.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj14.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj15.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj16.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj17.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj18.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj19.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj20.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj21.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj22.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(47, 7, '%s :' % (cell_obj23.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj24.value), ln=True)
pdf.ln(5)
pdf.cell(47, 7, '%s :' % (cell_object.value))
pdf.cell(20, 7, '%s' % (cell_objr.value), ln=True)
pdf.image('Student1.jpg', 10, 140, 80)

pdf.output('student_1.pdf')


path = "Dummy Data.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj1 = sheet_obj.cell(row=2, column=2)
cell_obj2 = sheet_obj.cell(row=28, column=2)
cell_obj3 = sheet_obj.cell(row=2, column=3)
cell_obj4 = sheet_obj.cell(row=28, column=3)
cell_obj5 = sheet_obj.cell(row=2, column=4)
cell_obj6 = sheet_obj.cell(row=28, column=4)
cell_obj7 = sheet_obj.cell(row=2, column=5)
cell_obj8 = sheet_obj.cell(row=28, column=5)
cell_obj9 = sheet_obj.cell(row=2, column=6)
cell_obj10 = sheet_obj.cell(row=28, column=6)
cell_obj11 = sheet_obj.cell(row=2, column=7)
cell_obj12 = sheet_obj.cell(row=28, column=7)
cell_obj13 = sheet_obj.cell(row=2, column=8)
cell_obj14 = sheet_obj.cell(row=28, column=8)
cell_obj15 = sheet_obj.cell(row=2, column=9)
cell_obj16 = sheet_obj.cell(row=28, column=9)
cell_obj17 = sheet_obj.cell(row=2, column=10)
cell_obj18 = sheet_obj.cell(row=28, column=10)
cell_obj19 = sheet_obj.cell(row=2, column=11)
cell_obj20 = sheet_obj.cell(row=28, column=11)
cell_obj21 = sheet_obj.cell(row=2, column=12)
cell_obj22 = sheet_obj.cell(row=28, column=12)
cell_obj23 = sheet_obj.cell(row=2, column=13)
cell_obj24 = sheet_obj.cell(row=28, column=13)
cell_object = sheet_obj.cell(row=2, column=20)
cell_objr = sheet_obj.cell(row=28, column=20)
# Print value of cell object
# using the value attribute


class PDF(FPDF):
    def header(self):
        # Logo
        self.image('Logo.jpg', 10, 8, 25)
        # font
        self.set_font('helvetica', 'I', 20)
        # Padding
        self.cell(60)
        # Title
        self.cell(100, 10, 'General Aptitude Examination',
                  border=True, ln=1, align='C')
        # Line break
        self.ln(20)


# create FPDF object
# Layout ('P','L')
# Unit ('mm', 'cm', 'in')
# format ('A3', 'A4' (default), 'A5', 'Letter', 'Legal', (100,150))
pdf = PDF('P', 'mm', 'Letter')

# Add a page
pdf.add_page()

# specify font
# fonts ('times', 'courier', 'helvetica', 'symbol', 'zpfdingbats')
# 'B' (bold), 'U' (underline), 'I' (italics), '' (regular), combination (i.e., ('BU'))
pdf.set_font('helvetica', 'B', 12)

# Add text
# w = width
# h = height
# txt = your text
# ln (0 False; 1 True - move cursor down to next line)
# border (0 False; 1 True - add border around cell)
pdf.cell(20, 7, '%s :' % (cell_obj1.value))
pdf.image('Pics for assignment/ABC2 XYZ2.png', 145, 40, 60)
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj2.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj3.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj5.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj7.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value) + ' %s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj9.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj10.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj11.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj12.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj13.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj14.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj15.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj16.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj17.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj18.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj19.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj20.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj21.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj22.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(47, 7, '%s :' % (cell_obj23.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj24.value), ln=True)
pdf.ln(5)
pdf.cell(47, 7, '%s :' % (cell_object.value))
pdf.cell(20, 7, '%s' % (cell_objr.value), ln=True)
pdf.image('Student2.jpg', 10, 140, 80)

pdf.output('student_2.pdf')


path = "Dummy Data.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj1 = sheet_obj.cell(row=2, column=2)
cell_obj2 = sheet_obj.cell(row=53, column=2)
cell_obj3 = sheet_obj.cell(row=2, column=3)
cell_obj4 = sheet_obj.cell(row=53, column=3)
cell_obj5 = sheet_obj.cell(row=2, column=4)
cell_obj6 = sheet_obj.cell(row=53, column=4)
cell_obj7 = sheet_obj.cell(row=2, column=5)
cell_obj8 = sheet_obj.cell(row=53, column=5)
cell_obj9 = sheet_obj.cell(row=2, column=6)
cell_obj10 = sheet_obj.cell(row=53, column=6)
cell_obj11 = sheet_obj.cell(row=2, column=7)
cell_obj12 = sheet_obj.cell(row=53, column=7)
cell_obj13 = sheet_obj.cell(row=2, column=8)
cell_obj14 = sheet_obj.cell(row=53, column=8)
cell_obj15 = sheet_obj.cell(row=2, column=9)
cell_obj16 = sheet_obj.cell(row=53, column=9)
cell_obj17 = sheet_obj.cell(row=2, column=10)
cell_obj18 = sheet_obj.cell(row=53, column=10)
cell_obj19 = sheet_obj.cell(row=2, column=11)
cell_obj20 = sheet_obj.cell(row=53, column=11)
cell_obj21 = sheet_obj.cell(row=2, column=12)
cell_obj22 = sheet_obj.cell(row=53, column=12)
cell_obj23 = sheet_obj.cell(row=2, column=13)
cell_obj24 = sheet_obj.cell(row=53, column=13)
cell_object = sheet_obj.cell(row=2, column=20)
cell_objr = sheet_obj.cell(row=53, column=20)
# Print value of cell object
# using the value attribute


class PDF(FPDF):
    def header(self):
        # Logo
        self.image('Logo.jpg', 10, 8, 25)
        # font
        self.set_font('helvetica', 'I', 20)
        # Padding
        self.cell(60)
        # Title
        self.cell(100, 10, 'General Aptitude Examination',
                  border=True, ln=1, align='C')
        # Line break
        self.ln(20)


# create FPDF object
# Layout ('P','L')
# Unit ('mm', 'cm', 'in')
# format ('A3', 'A4' (default), 'A5', 'Letter', 'Legal', (100,150))
pdf = PDF('P', 'mm', 'Letter')

# Add a page
pdf.add_page()

# specify font
# fonts ('times', 'courier', 'helvetica', 'symbol', 'zpfdingbats')
# 'B' (bold), 'U' (underline), 'I' (italics), '' (regular), combination (i.e., ('BU'))
pdf.set_font('helvetica', 'B', 12)

# Add text
# w = width
# h = height
# txt = your text
# ln (0 False; 1 True - move cursor down to next line)
# border (0 False; 1 True - add border around cell)
pdf.cell(20, 7, '%s :' % (cell_obj1.value))
pdf.image('Pics for assignment/ABC3 XYZ3.png', 145, 40, 50)
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj2.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj3.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj5.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj7.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value) + ' %s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj9.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj10.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj11.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj12.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj13.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj14.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj15.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj16.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj17.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj18.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj19.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj20.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj21.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj22.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(47, 7, '%s :' % (cell_obj23.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj24.value), ln=True)
pdf.ln(5)
pdf.cell(47, 7, '%s :' % (cell_object.value))
pdf.cell(20, 7, '%s' % (cell_objr.value), ln=True)
pdf.image('Student3.jpg', 10, 140, 80)

pdf.output('student_3.pdf')

path = "Dummy Data.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj1 = sheet_obj.cell(row=2, column=2)
cell_obj2 = sheet_obj.cell(row=78, column=2)
cell_obj3 = sheet_obj.cell(row=2, column=3)
cell_obj4 = sheet_obj.cell(row=78, column=3)
cell_obj5 = sheet_obj.cell(row=2, column=4)
cell_obj6 = sheet_obj.cell(row=78, column=4)
cell_obj7 = sheet_obj.cell(row=2, column=5)
cell_obj8 = sheet_obj.cell(row=78, column=5)
cell_obj9 = sheet_obj.cell(row=2, column=6)
cell_obj10 = sheet_obj.cell(row=78, column=6)
cell_obj11 = sheet_obj.cell(row=2, column=7)
cell_obj12 = sheet_obj.cell(row=78, column=7)
cell_obj13 = sheet_obj.cell(row=2, column=8)
cell_obj14 = sheet_obj.cell(row=78, column=8)
cell_obj15 = sheet_obj.cell(row=2, column=9)
cell_obj16 = sheet_obj.cell(row=78, column=9)
cell_obj17 = sheet_obj.cell(row=2, column=10)
cell_obj18 = sheet_obj.cell(row=78, column=10)
cell_obj19 = sheet_obj.cell(row=2, column=11)
cell_obj20 = sheet_obj.cell(row=78, column=11)
cell_obj21 = sheet_obj.cell(row=2, column=12)
cell_obj22 = sheet_obj.cell(row=78, column=12)
cell_obj23 = sheet_obj.cell(row=2, column=13)
cell_obj24 = sheet_obj.cell(row=78, column=13)
cell_object = sheet_obj.cell(row=2, column=20)
cell_objr = sheet_obj.cell(row=78, column=20)
# Print value of cell object
# using the value attribute


class PDF(FPDF):
    def header(self):
        # Logo
        self.image('Logo.jpg', 10, 8, 25)
        # font
        self.set_font('helvetica', 'I', 20)
        # Padding
        self.cell(60)
        # Title
        self.cell(100, 10, 'General Aptitude Examination',
                  border=True, ln=1, align='C')
        # Line break
        self.ln(20)


# create FPDF object
# Layout ('P','L')
# Unit ('mm', 'cm', 'in')
# format ('A3', 'A4' (default), 'A5', 'Letter', 'Legal', (100,150))
pdf = PDF('P', 'mm', 'Letter')

# Add a page
pdf.add_page()

# specify font
# fonts ('times', 'courier', 'helvetica', 'symbol', 'zpfdingbats')
# 'B' (bold), 'U' (underline), 'I' (italics), '' (regular), combination (i.e., ('BU'))
pdf.set_font('helvetica', 'B', 12)

# Add text
# w = width
# h = height
# txt = your text
# ln (0 False; 1 True - move cursor down to next line)
# border (0 False; 1 True - add border around cell)
pdf.cell(20, 7, '%s :' % (cell_obj1.value))
pdf.image('Pics for assignment/ABC4 XYZ4.png', 145, 40, 60)
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj2.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj3.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj5.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj7.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value) + ' %s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj9.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj10.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj11.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj12.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj13.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj14.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj15.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj16.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj17.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj18.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj19.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj20.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj21.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj22.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(47, 7, '%s :' % (cell_obj23.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj24.value), ln=True)
pdf.ln(5)
pdf.cell(47, 7, '%s :' % (cell_object.value))
pdf.cell(20, 7, '%s' % (cell_objr.value), ln=True)
pdf.image('Student4.jpg', 10, 140, 80)

pdf.output('student_4.pdf')

path = "Dummy Data.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj1 = sheet_obj.cell(row=2, column=2)
cell_obj2 = sheet_obj.cell(row=103, column=2)
cell_obj3 = sheet_obj.cell(row=2, column=3)
cell_obj4 = sheet_obj.cell(row=103, column=3)
cell_obj5 = sheet_obj.cell(row=2, column=4)
cell_obj6 = sheet_obj.cell(row=103, column=4)
cell_obj7 = sheet_obj.cell(row=2, column=5)
cell_obj8 = sheet_obj.cell(row=103, column=5)
cell_obj9 = sheet_obj.cell(row=2, column=6)
cell_obj10 = sheet_obj.cell(row=103, column=6)
cell_obj11 = sheet_obj.cell(row=2, column=7)
cell_obj12 = sheet_obj.cell(row=103, column=7)
cell_obj13 = sheet_obj.cell(row=2, column=8)
cell_obj14 = sheet_obj.cell(row=103, column=8)
cell_obj15 = sheet_obj.cell(row=2, column=9)
cell_obj16 = sheet_obj.cell(row=103, column=9)
cell_obj17 = sheet_obj.cell(row=2, column=10)
cell_obj18 = sheet_obj.cell(row=103, column=10)
cell_obj19 = sheet_obj.cell(row=2, column=11)
cell_obj20 = sheet_obj.cell(row=103, column=11)
cell_obj21 = sheet_obj.cell(row=2, column=12)
cell_obj22 = sheet_obj.cell(row=103, column=12)
cell_obj23 = sheet_obj.cell(row=2, column=13)
cell_obj24 = sheet_obj.cell(row=103, column=13)
cell_object = sheet_obj.cell(row=2, column=20)
cell_objr = sheet_obj.cell(row=103, column=20)
# Print value of cell object
# using the value attribute


class PDF(FPDF):
    def header(self):
        # Logo
        self.image('Logo.jpg', 10, 8, 25)
        # font
        self.set_font('helvetica', 'I', 20)
        # Padding
        self.cell(60)
        # Title
        self.cell(100, 10, 'General Aptitude Examination',
                  border=True, ln=1, align='C')
        # Line break
        self.ln(20)


# create FPDF object
# Layout ('P','L')
# Unit ('mm', 'cm', 'in')
# format ('A3', 'A4' (default), 'A5', 'Letter', 'Legal', (100,150))
pdf = PDF('P', 'mm', 'Letter')

# Add a page
pdf.add_page()

# specify font
# fonts ('times', 'courier', 'helvetica', 'symbol', 'zpfdingbats')
# 'B' (bold), 'U' (underline), 'I' (italics), '' (regular), combination (i.e., ('BU'))
pdf.set_font('helvetica', 'B', 12)

# Add text
# w = width
# h = height
# txt = your text
# ln (0 False; 1 True - move cursor down to next line)
# border (0 False; 1 True - add border around cell)
pdf.cell(20, 7, '%s :' % (cell_obj1.value))
pdf.image('Pics for assignment/ABC5 XYZ5.png', 145, 40, 60)
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj2.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj3.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj5.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj7.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj4.value) + ' %s' % (cell_obj6.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj9.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj10.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj11.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj12.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj13.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj14.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj15.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj16.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(30, 7, '%s :' % (cell_obj17.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj18.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(40, 7, '%s :' % (cell_obj19.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj20.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(45, 7, '%s :' % (cell_obj21.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj22.value), ln=True)
pdf.set_font('helvetica', 'B', 12)
pdf.cell(47, 7, '%s :' % (cell_obj23.value))
pdf.set_font('helvetica', '', 12)
pdf.cell(20, 7, '%s' % (cell_obj24.value), ln=True)
pdf.ln(5)
pdf.cell(47, 7, '%s :' % (cell_object.value))
pdf.cell(20, 7, '%s' % (cell_objr.value), ln=True)
pdf.image('Student5.jpg', 10, 140, 80)

pdf.output('student_5.pdf')
